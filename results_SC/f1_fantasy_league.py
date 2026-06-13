"""
F1 Fantasy League — Full Pipeline
===================================
What this script does in one run:
  1. Fetches full league data from F1 Fantasy API
  2. Recalculates LIVE rankings from gd_points (accurate during race weekend)
  3. Saves a per-race JSON + Excel file to a local folder (syncs to Google Drive)
  4. Updates Google Sheets (duplicate-safe — safe to run multiple times per race)
       - "Current Race" tab: always overwritten with latest data
       - "History" tab: replaces rows for current race, appends new ones
       - "Picks Detail" tab: same logic as History
  5. Sends WhatsApp message with live standings

SETUP INSTRUCTIONS (one time only):
─────────────────────────────────────────────────────────────────────

1. INSTALL DEPENDENCIES:
     pip install requests openpyxl gspread oauth2client

2. SET YOUR COOKIE (refresh when you get 401 errors):
     - Go to fantasy.formula1.com (logged in)
     - DevTools (F12) → Network → click any /services/ request
     - Right-click → Copy → Copy as cURL
     - Paste in Notepad, find: -H 'cookie: XXXX'
     - Copy just that cookie value (everything after 'cookie: ')
     - Paste it into cookie.txt (same folder as this script)
     - No quotes, no extra spaces — just the raw cookie string

3. GOOGLE SHEETS SETUP (one time):
     a. Go to console.cloud.google.com
     b. Create project → Enable Google Sheets API + Google Drive API
     c. APIs & Services → Credentials → Create Service Account → name it "f1-fantasy-bot"
     d. Click the service account → Keys tab → Add Key → JSON → Download
     e. Rename downloaded file to "google_credentials.json"
     f. Move it to the SAME folder as this script
     g. Open the JSON, copy the "client_email" value
     h. Create a Google Sheet named "F1 Fantasy Dashboard"
     i. Share that sheet with the client_email as Editor
     j. Set ENABLE_SHEETS = True below

4. CALLMEBOT WHATSAPP (one time):
     - Send "I allow callmebot to send me messages" to +34 644 59 78 70 on WhatsApp
     - You'll receive an API key
     - Set ENABLE_WHATSAPP = True and fill WHATSAPP_RECIPIENTS below

─────────────────────────────────────────────────────────────────────
"""

import requests
import json
import time
import os
import re
from datetime import datetime, timezone
from urllib.parse import unquote

from f1_config import cfg, SCRIPT_DIR
from f1_calendar import fetch_calendar, detect_current_race, detect_last_completed_race

# ══════════════════════════════════════════════════════════════════
#  SETTINGS — loaded from config.json via f1_config.py
#  No manual changes needed here between races!
# ══════════════════════════════════════════════════════════════════

USER_UUID   = cfg.get("user_uuid", "")
LEAGUE_ID   = cfg.get("league_id", "")
LEAGUE_NAME = cfg.get("league_name", "F1 Fantasy")
SAVE_FOLDER = cfg.get("save_folder", ".")

# Google Sheets
ENABLE_SHEETS     = cfg.get("enable_sheets", False)
GOOGLE_CREDS_FILE = cfg.get("google_creds_file", os.path.join(SCRIPT_DIR, "google_credentials.json"))
SHEET_NAME        = cfg.get("sheet_name", "F1 Fantasy Dashboard")

# WhatsApp
ENABLE_WHATSAPP     = cfg.get("enable_whatsapp", False)
WHATSAPP_RECIPIENTS = [(r["phone"], r["apikey"]) for r in cfg.get("whatsapp", [])]

# ── Cookie ────────────────────────────────────────────────────────
COOKIE_FILE = os.path.join(SCRIPT_DIR, "cookie.txt")

def load_cookie():
    if not os.path.exists(COOKIE_FILE):
        print(f"\n⚠️  ERROR: cookie.txt not found!")
        print(f"  Create a file named 'cookie.txt' in: {SCRIPT_DIR}")
        print(f"  Paste your F1 Fantasy cookie value into it and save.")
        return None
    cookie = open(COOKIE_FILE, encoding="utf-8").read().strip()
    if not cookie:
        print(f"\n⚠️  ERROR: cookie.txt is empty!")
        print(f"  Paste your F1 Fantasy cookie value into cookie.txt and save.")
        return None
    return cookie

COOKIE = load_cookie() or ""

# ── Auto-detect race from calendar ───────────────────────────────
def _detect_race():
    """Try to auto-detect current race from API. Falls back to config.json."""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept":     "application/json, text/plain, */*",
            "Referer":    "https://fantasy.formula1.com/",
            "Cookie":     COOKIE,
        }
        calendar     = fetch_calendar(headers)
        current_race = detect_last_completed_race(calendar)

        # ── --matchday override (set by run_all.py --matchday N) ──
        override = cfg.get("matchday_override")
        if override:
            match = next((r for r in calendar if r["matchday"] == override), None)
            if match:
                print(f"  ⚡ Matchday override: Race {override} — {match['race_name']}")
                return (
                    match["matchday"],
                    match["race_name"],
                    match.get("race_flag", "🏎️"),
                )
            else:
                print(f"  ⚠  Override Race {override} not found in calendar — falling back")

        if current_race:
            print(f"  ✅ Auto-detected: Race {current_race['matchday']} — {current_race['race_name']}")
            return (
                current_race["matchday"],
                current_race["race_name"],
                current_race.get("race_flag", "🏎️"),
            )
    except Exception as e:
        print(f"  ⚠  Calendar auto-detect failed: {e}")

    # Fallback to config.json manual override
    matchday  = cfg.get("matchday", 1)
    race_name = cfg.get("race_name", "Race")
    race_flag = cfg.get("race_flag", "🏎️")
    print(f"  ⚠  Using config.json fallback: Race {matchday} — {race_name}")
    return matchday, race_name, race_flag

MATCHDAY, RACE_NAME, RACE_FLAG = _detect_race()

# ══════════════════════════════════════════════════════════════════

BASE_SVC  = "https://fantasy.formula1.com/services/user"
BASE_FEED = "https://fantasy.formula1.com/feeds"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept":     "application/json, text/plain, */*",
    "Referer":    "https://fantasy.formula1.com/",
    "Cookie":     COOKIE,
}

DRIVER_MAP = {
    "11":    "Alexander Albon",    # Williams
    "11149": "Arvid Lindblad",     # Racing Bulls
    "125":   "Carlos Sainz",       # Williams
    "115":   "Charles Leclerc",    # Ferrari
    "118":   "Esteban Ocon",       # Haas
    "12":    "Fernando Alonso",    # Aston Martin
    "11059": "Franco Colapinto",   # Alpine
    "11051": "Gabriel Bortoleto",  # Audi
    "124":   "George Russell",     # Mercedes
    "11032": "Isack Hadjar",       # Red Bull Racing
    "11161": "Kimi Antonelli",     # Mercedes
    "129":   "Lance Stroll",       # Aston Martin
    "117":   "Lando Norris",       # McLaren
    "110":   "Lewis Hamilton",     # Ferrari
    "114":   "Liam Lawson",        # Racing Bulls
    "131":   "Max Verstappen",     # Red Bull Racing
    "111":   "Nico Hulkenberg",    # Audi
    "11031": "Oliver Bearman",     # Haas
    "1982":  "Oscar Piastri",      # McLaren
    "18":    "Pierre Gasly",       # Alpine
    "121":   "Sergio Perez",       # Cadillac
    "13":    "Valtteri Bottas",    # Cadillac
}

CONSTRUCTOR_MAP = {
    "23":   "Alpine",
    "24":   "Aston Martin",
    "2640": "Audi",
    "2641": "Cadillac",
    "25":   "Ferrari",
    "26":   "Haas F1 Team",
    "27":   "McLaren",
    "28":   "Mercedes",
    "2636": "Racing Bulls",
    "29":   "Red Bull Racing",
    "210":  "Williams",
}

CARD_FIELDS = {
    "isWildcardtaken":   ("Wildcard",    "wildCardtakengd"),
    "isLimitlesstaken":  ("Limitless",   "limitLesstakengd"),
    "isNonigativetaken": ("No Negative", "noNigativetakengd"),
    "isFinalfixtaken":   ("Final Fix",   "finalFixtakengd"),
    "isExtradrstaken":   ("Extra DRS",   "extraDrstakengd"),
    "isAutopilottaken":  ("Autopilot",   "isAutopilottakengd"),
}

RANK_EMOJI = {
    1: "1⃣", 2: "2⃣", 3: "3⃣",
    4: "4⃣", 5: "5⃣", 6: "6⃣",
    7: "7⃣", 8: "8⃣", 9: "9⃣", 10: "🔟",
    11: "11.", 12: "12.", 13: "13."
}


# ── Helpers ───────────────────────────────────────────────────────

def get(url):
    r = requests.get(url, headers=HEADERS, timeout=15)
    r.raise_for_status()
    return r.json()

def buster():
    return int(time.time() * 1000)

def slugify(name):
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")

def rank_change_arrow(change):
    if change > 0:   return f"▲{change}"
    elif change < 0: return f"▼{abs(change)}"
    return "─"


# ── Rank recalculation ────────────────────────────────────────────

def recalculate_ranks(results):
    """
    Recalculates overall rank (from live total_points) and
    gameday rank (from live gd_points). Handles ties correctly.
    Also computes rank change vs cached API rank.
    """
    def assign_ranks(items, key):
        sorted_items = sorted(items, key=lambda x: x[key] if x[key] is not None else 0, reverse=True)
        rank = 1
        for i, t in enumerate(sorted_items):
            if i > 0 and t[key] == sorted_items[i-1][key]:
                t[f"_{key}_rank"] = sorted_items[i-1][f"_{key}_rank"]
            else:
                t[f"_{key}_rank"] = rank
            rank += 1

    assign_ranks(results, "total_points")
    assign_ranks(results, "gd_points")

    for t in results:
        t["overall_rank"]  = t.pop("_total_points_rank")
        t["live_gd_rank"]  = t.pop("_gd_points_rank")
        t["rank_change"]   = t["rank"] - t["overall_rank"]

    return sorted(results, key=lambda x: (x["overall_rank"], x["live_gd_rank"]))


# ── API calls ─────────────────────────────────────────────────────

def get_live_buster():
    data = get(f"{BASE_FEED}/live/mixapi.json?buster={buster()}")
    return data["Value"]["lv"]

def get_leaderboard():
    url = f"{BASE_FEED}/leaderboard/privateleague/list_1_{LEAGUE_ID}_0_1.json?buster={buster()}"
    return get(url)["Value"]["leaderboard"]

def get_cards(user_guid):
    # Last param must be 1 (cumulative view) — passing MATCHDAY returns wrong card state.
    # Filter by takengd == MATCHDAY so only cards used THIS race are shown.
    url = f"{BASE_SVC}/opponentteam/opponentgamedayget/1/{user_guid}/1?buster={buster()}"
    data = get(url)["Data"]["Value"]
    cards = []
    for field, (label, gd_field) in CARD_FIELDS.items():
        if data.get(field, 0):
            gd = data.get(gd_field, "?")
            if gd == MATCHDAY:          # only include cards used in the current race
                cards.append(f"{label} (GD{gd})")
    return cards

def get_team_detail(user_guid):
    url = (f"{BASE_SVC}/opponentteam/opponentgamedayplayerteamget"
           f"/1/{user_guid}/1/{MATCHDAY}/1?buster={buster()}")
    return get(url)["Data"]["Value"]["userTeam"][0]

def get_player_stats(player_id, lv):
    url = f"{BASE_FEED}/popup/playerstats_{player_id}.json?buster={lv}"
    try:
        data = get(url)["Value"]["GamedayWiseStats"]
        for gd in data:
            if gd["GamedayId"] == MATCHDAY:
                total = next((s["Value"] for s in gd["StatsWise"]
                              if s["Event"] == "Total"), None)
                breakdown = [f"{s['Event']}: {s['Value']:+d}"
                             for s in gd["StatsWise"] if s["Event"] != "Total"]
                return total, breakdown
    except Exception:
        pass
    return None, []


# ── Step 1: Fetch all data ────────────────────────────────────────

def fetch_all():
    print("\n[1/5] Fetching live buster...")
    lv = get_live_buster()

    print("[2/5] Fetching leaderboard...")
    leaderboard = get_leaderboard()
    print(f"  Found {len(leaderboard)} teams")

    results = []
    print("[3/5] Fetching team details + player stats...")

    for entry in leaderboard:
        user_guid  = entry["user_guid"]
        team_name  = unquote(entry["team_name"])
        user_name  = entry["user_name"]
        api_rank   = int(entry["cur_rank"])
        cur_points = entry["cur_points"]

        print(f"  → {team_name} ({user_name})")

        try:
            cards = get_cards(user_guid)
        except Exception:
            cards = []

        try:
            detail     = get_team_detail(user_guid)
            budget     = detail["team_info"]["teamBal"]
            team_value = detail["team_info"]["teamVal"]
            gd_points  = detail["gdpoints"] or 0
            ovpoints   = detail["ovpoints"] or 0
            player_ids = detail["playerid"]
        except Exception as e:
            print(f"    ⚠ Team detail error: {e}")
            continue

        picks = []
        for pick in sorted(player_ids, key=lambda x: x["playerpostion"]):
            pid      = pick["id"]
            is_cap   = pick["iscaptain"] == 1
            is_mgcap = pick["ismgcaptain"] == 1
            ptype    = "Constructor" if pid in CONSTRUCTOR_MAP else "Driver"
            name     = CONSTRUCTOR_MAP.get(pid) or DRIVER_MAP.get(pid) or f"Unknown (ID {pid})"
            pts, breakdown = get_player_stats(pid, lv)
            picks.append({
                "id":         pid,
                "name":       name,
                "type":       ptype,
                "is_captain": is_cap,
                "is_megacap": is_mgcap,
                "points_gd":  pts,
                "breakdown":  breakdown,
            })
            time.sleep(0.1)

        results.append({
            "race_number":      MATCHDAY,
            "race_name":        RACE_NAME,
            "race_flag":        RACE_FLAG,
            "race_date":        datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "rank":             api_rank,
            "overall_rank":     0,
            "live_gd_rank":     0,
            "rank_change":      0,
            "team_name":        team_name,
            "user_name":        user_name,
            "total_points":     ovpoints,
            "gd_points":        gd_points,
            "budget_remaining": budget,
            "team_value":       team_value,
            "cards_used":       cards,
            "picks":            picks,
        })

    print("[4/5] Recalculating live rankings...")
    results = recalculate_ranks(results)
    for t in results:
        arrow = rank_change_arrow(t["rank_change"])
        print(f"  #{t['overall_rank']} {t['team_name']} — "
              f"{t['total_points']}pts total | {t['gd_points']}pts this race {arrow}")

    return results


# ── Step 2: Save Excel ────────────────────────────────────────────

def save_files(results):
    os.makedirs(SAVE_FOLDER, exist_ok=True)
    slug = f"race_{MATCHDAY:02d}_{slugify(RACE_NAME)}"

    # JSON backup
    json_path = os.path.join(SAVE_FOLDER, f"{slug}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"  ✅ JSON: {json_path}")

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb       = openpyxl.Workbook()
        hdr_fill = PatternFill("solid", fgColor="E8041D")
        hdr_font = Font(bold=True, color="FFFFFF")

        def style_header(ws, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font      = hdr_font
                c.fill      = hdr_fill
                c.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"

        # Sheet 1: Standings
        ws1 = wb.active
        ws1.title = "Standings"
        style_header(ws1, [
            "Overall Rank", "GD Rank", "Team Name", "Owner",
            "Total Points", f"GD{MATCHDAY} Points",
            "Budget ($m)", "Team Value ($m)", "Cards Used", "Change"
        ])
        for row, t in enumerate(results, 2):
            ws1.cell(row=row, column=1,  value=t["overall_rank"])
            ws1.cell(row=row, column=2,  value=t["live_gd_rank"])
            ws1.cell(row=row, column=3,  value=t["team_name"])
            ws1.cell(row=row, column=4,  value=t["user_name"])
            ws1.cell(row=row, column=5,  value=t["total_points"])
            ws1.cell(row=row, column=6,  value=t["gd_points"])
            ws1.cell(row=row, column=7,  value=t["budget_remaining"])
            ws1.cell(row=row, column=8,  value=t["team_value"])
            ws1.cell(row=row, column=9,  value=", ".join(t["cards_used"]) or "None")
            ws1.cell(row=row, column=10, value=rank_change_arrow(t["rank_change"]))
        for col in range(1, 11):
            ws1.column_dimensions[get_column_letter(col)].width = 20

        # Sheet 2: Picks Detail
        ws2 = wb.create_sheet("Picks Detail")
        style_header(ws2, [
            "Race", "Overall Rank", "Team", "Owner",
            "Type", "Player / Constructor", "Captain",
            "GD Points", "Breakdown"
        ])
        row = 2
        for t in results:
            for p in t["picks"]:
                ws2.cell(row=row, column=1, value=RACE_NAME)
                ws2.cell(row=row, column=2, value=t["overall_rank"])
                ws2.cell(row=row, column=3, value=t["team_name"])
                ws2.cell(row=row, column=4, value=t["user_name"])
                ws2.cell(row=row, column=5, value=p["type"])
                ws2.cell(row=row, column=6, value=p["name"])
                cap = "★ Captain" if p["is_captain"] else ("★ Mega" if p["is_megacap"] else "")
                ws2.cell(row=row, column=7, value=cap)
                ws2.cell(row=row, column=8, value=p["points_gd"])
                ws2.cell(row=row, column=9, value=" | ".join(p["breakdown"]))
                row += 1
        for col in range(1, 10):
            ws2.column_dimensions[get_column_letter(col)].width = 22

        xlsx_path = os.path.join(SAVE_FOLDER, f"{slug}.xlsx")
        wb.save(xlsx_path)
        print(f"  ✅ Excel: {xlsx_path}")

    except ImportError:
        print("  ⚠ openpyxl not installed — run: pip install openpyxl")
    except PermissionError:
        print(f"  ⚠ Excel file is open — close it and run again")


# ── Step 3: Google Sheets (duplicate-safe) ────────────────────────

def update_google_sheet(results):
    if not ENABLE_SHEETS:
        print("  ⏭ Google Sheets skipped (set ENABLE_SHEETS = True to enable)")
        return

    try:
        import gspread
        from oauth2client.service_account import ServiceAccountCredentials

        scope  = ["https://spreadsheets.google.com/feeds",
                  "https://www.googleapis.com/auth/drive"]
        creds  = ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_CREDS_FILE, scope)
        client = gspread.authorize(creds)
        sheet  = client.open(SHEET_NAME)

        # ── Tab 1: Current Race (always fully overwritten) ──────
        try:
            ws_current = sheet.worksheet("Current Race")
            ws_current.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws_current = sheet.add_worksheet("Current Race", rows=50, cols=12)

        ws_current.update("A1", [[
            f"{RACE_FLAG} {RACE_NAME} — Race {MATCHDAY}  |  "
            f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC"
        ]])
        ws_current.update("A2", [[
            "Overall Rank", "GD Rank", "Team Name", "Owner",
            "Total Points", f"GD{MATCHDAY} Points",
            "Budget ($m)", "Team Value ($m)", "Cards Used", "Change"
        ]])
        current_rows = []
        for t in results:
            current_rows.append([
                t["overall_rank"], t["live_gd_rank"],
                t["team_name"],    t["user_name"],
                t["total_points"], t["gd_points"],
                t["budget_remaining"], t["team_value"],
                ", ".join(t["cards_used"]) or "None",
                rank_change_arrow(t["rank_change"])
            ])
        ws_current.update("A3", current_rows)
        print(f"  ✅ 'Current Race' tab updated")

        # ── Tab 2: History (remove this race's rows, re-append) ─
        try:
            ws_hist = sheet.worksheet("History")
        except gspread.exceptions.WorksheetNotFound:
            ws_hist = sheet.add_worksheet("History", rows=1000, cols=11)
            ws_hist.update("A1", [[
                "Race #", "Race Name", "Date", "Overall Rank", "GD Rank",
                "Team", "Owner", "Total Pts", "GD Pts", "Budget ($m)", "Cards Used"
            ]])

        # Delete existing rows for this MATCHDAY (avoid duplicates)
        all_hist = ws_hist.get_all_values()
        rows_to_delete = [
            i + 1 for i, row in enumerate(all_hist)
            if i > 0 and len(row) > 0 and str(row[0]) == str(MATCHDAY)
        ]
        # Delete from bottom up to preserve row indices
        for row_idx in reversed(rows_to_delete):
            ws_hist.delete_rows(row_idx)
            time.sleep(0.3)  # avoid API rate limit

        # Append fresh rows for this race
        new_hist_rows = []
        for t in results:
            new_hist_rows.append([
                t["race_number"], t["race_name"], t["race_date"],
                t["overall_rank"], t["live_gd_rank"],
                t["team_name"],   t["user_name"],
                t["total_points"], t["gd_points"],
                t["budget_remaining"],
                ", ".join(t["cards_used"]) or "None"
            ])
        ws_hist.append_rows(new_hist_rows)
        print(f"  ✅ 'History' tab updated ({len(rows_to_delete)} old rows replaced)")

        # ── Tab 3: Picks Detail (remove this race, re-append) ───
        try:
            ws_picks = sheet.worksheet("Picks Detail")
        except gspread.exceptions.WorksheetNotFound:
            ws_picks = sheet.add_worksheet("Picks Detail", rows=2000, cols=10)
            ws_picks.update("A1", [[
                "Race #", "Race Name", "Overall Rank", "Team", "Owner",
                "Type", "Player / Constructor", "Captain", "GD Points", "Breakdown"
            ]])

        # Delete existing rows for this MATCHDAY
        all_picks = ws_picks.get_all_values()
        pick_rows_to_delete = [
            i + 1 for i, row in enumerate(all_picks)
            if i > 0 and len(row) > 0 and str(row[0]) == str(MATCHDAY)
        ]
        for row_idx in reversed(pick_rows_to_delete):
            ws_picks.delete_rows(row_idx)
            time.sleep(0.3)

        # Append fresh pick rows
        new_pick_rows = []
        for t in results:
            for p in t["picks"]:
                cap = "★ Captain" if p["is_captain"] else ("★ Mega" if p["is_megacap"] else "")
                new_pick_rows.append([
                    t["race_number"], t["race_name"],
                    t["overall_rank"], t["team_name"], t["user_name"],
                    p["type"], p["name"], cap,
                    p["points_gd"],
                    " | ".join(p["breakdown"])
                ])
        ws_picks.append_rows(new_pick_rows)
        print(f"  ✅ 'Picks Detail' tab updated ({len(pick_rows_to_delete)} old rows replaced)")

        # ── Tab 4: Season Standings (full rebuild each run) ────────
        PLAYER_CODES = {
            "Danilo Iglesias":       "DAN",
            "Eduardo Santos Lima":   "EDU",
            "Fabio Mucci":           "FAB",
            "Guilherme Figueiredo":  "GUI",
            "Liberio Junior":        "LIB",
            "Priscila Rigotto":      "PRI",
            "Rafael Dias":           "RAF",
            "Raphael Stein":         "STE",
            "Ricardo Mucci":         "RIC",
            "Robson Jardim":         "ROB",
            "Vinicius Agostini":     "VIN",
        }

        try:
            ws_season = sheet.worksheet("Season Standings")
            ws_season.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws_season = sheet.add_worksheet("Season Standings", rows=50, cols=30)

        # Read full history to build season matrix
        all_hist = ws_hist.get_all_values()
        if len(all_hist) > 1:
            # Parse history rows: Race#, Race Name, Date, Overall Rank, GD Rank,
            #                     Team, Owner, Total Pts, GD Pts, Budget, Cards
            from collections import defaultdict
            # race_data[race_num] = {user_name: {gd_pts, total_pts, rank, flag, name}}
            race_data   = defaultdict(dict)
            race_meta   = {}   # race_num -> (race_name, race_flag)
            user_totals = {}   # user_name -> latest total_pts

            for row in all_hist[1:]:
                if len(row) < 9:
                    continue
                try:
                    rn        = int(row[0])
                    race_name = row[1]
                    ovrank    = int(row[3])
                    gd_rank   = int(row[4])
                    user_name = row[6]
                    total_pts = float(row[7])
                    gd_pts    = float(row[8])
                except (ValueError, IndexError):
                    continue

                # Flag: look up from current results if this race matches, else blank
                flag = ""
                for r in results:
                    if r["race_number"] == rn:
                        flag = r.get("race_flag", "")
                        break

                race_meta[rn] = (race_name, flag)
                race_data[rn][user_name] = {
                    "gd_pts":    gd_pts,
                    "total_pts": total_pts,
                    "ov_rank":   ovrank,
                }
                # Keep latest total_pts per user (highest race number = most recent)
                if user_name not in user_totals:
                    user_totals[user_name] = (rn, total_pts)
                elif rn >= user_totals[user_name][0]:
                    user_totals[user_name] = (rn, total_pts)

            # All races sorted
            all_races    = sorted(race_meta.keys())
            all_users    = sorted(user_totals.keys(),
                                  key=lambda u: -user_totals[u][1])

            # Build header row: Player | R1 🇦🇺 | R2 🇨🇳 | ... | Total | Trend
            def race_header(rn):
                name, flag = race_meta[rn]
                # Short name: first word or abbreviation
                short = name.replace(" Grand Prix", "").replace(" GP", "").strip()
                short = short[:6] if len(short) > 6 else short
                return f"R{rn} {flag}" if flag else f"R{rn} {short}"

            header = ["Player"] + [race_header(rn) for rn in all_races] + ["Total", "Trend"]
            season_rows = [header]

            # One row per player
            for user_name in all_users:
                code = PLAYER_CODES.get(user_name, user_name[:3].upper())
                row_vals = [code]
                for rn in all_races:
                    entry = race_data[rn].get(user_name)
                    if entry:
                        # Format: #rank · pts
                        row_vals.append(f"#{entry['ov_rank']} · {int(entry['gd_pts'])}pts")
                    else:
                        row_vals.append("—")
                row_vals.append(int(user_totals[user_name][1]))  # Total
                row_vals.append("")  # Trend placeholder — sparkline added below
                season_rows.append(row_vals)

            ws_season.update("A1", season_rows, value_input_option="USER_ENTERED")

            # Add SPARKLINE formulas for Trend column
            # Column = 2 + len(all_races) + 1 (sparkline is last col, header row is row 1)
            # Data cols = B through (B + len(races) - 1), numeric only doesn't work with text
            # Instead use Total column as single-bar — real sparkline needs numeric per race
            # We'll add a note column with rank progression as numbers separately
            # For now: write rank numbers in a hidden numeric block, sparkline refs those

            # Write numeric GD pts in cols after Trend (hidden data for sparkline)
            hidden_start_col = len(all_races) + 3  # after Player, races, Total, Trend
            hidden_rows = []
            for user_name in all_users:
                numeric_row = []
                for rn in all_races:
                    entry = race_data[rn].get(user_name)
                    numeric_row.append(int(entry["gd_pts"]) if entry else 0)
                hidden_rows.append(numeric_row)

            # gspread col index helper
            from gspread.utils import rowcol_to_a1
            def col_letter(n):
                result = ""
                while n > 0:
                    n, r = divmod(n - 1, 26)
                    result = chr(65 + r) + result
                return result

            hidden_col_start = hidden_start_col
            hidden_col_end   = hidden_start_col + len(all_races) - 1
            hidden_start_a1  = f"{col_letter(hidden_col_start)}2"
            ws_season.update(hidden_start_a1, hidden_rows, value_input_option="USER_ENTERED")

            # Add SPARKLINE formula for each player row
            trend_col = col_letter(len(all_races) + 3)  # Trend column letter
            for row_i, user_name in enumerate(all_users):
                data_range = (
                    f"{col_letter(hidden_col_start)}{row_i+2}:"
                    f"{col_letter(hidden_col_end)}{row_i+2}"
                )
                formula = f'=SPARKLINE({data_range},{{"charttype","line";"color","#00D2B9";"linewidth",2}})' 
                ws_season.update_cell(row_i + 2, len(all_races) + 3, formula)
                time.sleep(0.15)

            # Hide the numeric helper columns (gspread can't hide cols, user does it manually)
            print(f"  ✅ 'Season Standings' tab updated ({len(all_users)} players, {len(all_races)} races)")
            print(f"     Tip: hide columns {col_letter(hidden_col_start)}–{col_letter(hidden_col_end)} in Sheets (they are sparkline data)")
        else:
            print(f"  ⚠  'Season Standings' skipped — History tab is empty")

    except FileNotFoundError:
        print(f"  ⚠ '{GOOGLE_CREDS_FILE}' not found — place it in the same folder as this script")
    except Exception as e:
        print(f"  ⚠ Google Sheets error: {e}")


# ── Step 4: WhatsApp ──────────────────────────────────────────────

def build_whatsapp_message(results):
    # Sort by this race's gd_points for the WhatsApp message
    sorted_results = sorted(results, key=lambda x: x["live_gd_rank"])
    date_str = datetime.now(timezone.utc).strftime("%d %b %Y")

    lines = [
        f"🏎️ {RACE_FLAG} {RACE_NAME} - Race {MATCHDAY}",
        f"📅 {date_str}",
        f"📊 {LEAGUE_NAME}",
        "",
    ]

    for t in sorted_results:
        emoji = RANK_EMOJI.get(t["live_gd_rank"], f"{t['live_gd_rank']}.")
        lines.append(f"{emoji} {t['team_name']} - {int(t['gd_points'])}pts")

    # Cards used — compact
    cards_lines = [
        f"  {t['user_name']}: {', '.join(t['cards_used'])}"
        for t in sorted_results if t["cards_used"]
    ]
    if cards_lines:
        lines += ["", "🃏 Cards:"] + cards_lines

    lines += ["", "🏁 Next race soon!"]
    return "\n".join(lines)

def send_whatsapp(results):
    message = build_whatsapp_message(results)

    if not ENABLE_WHATSAPP:
        print("  ⏭ WhatsApp skipped (set ENABLE_WHATSAPP = True to enable)")
        print("\n  📋 MESSAGE PREVIEW:")
        print("  " + "─" * 50)
        for line in message.split("\n"):
            print(f"  {line}")
        print("  " + "─" * 50)
        return

    for phone, api_key in WHATSAPP_RECIPIENTS:
        try:
            r = requests.get(
                "https://api.callmebot.com/whatsapp.php",
                params={"phone": phone, "text": message, "apikey": api_key},
                timeout=15
            )
            if r.status_code == 200:
                print(f"  ✅ WhatsApp sent to {phone}")
            else:
                print(f"  ⚠ Failed for {phone}: {r.text[:100]}")
            time.sleep(2)
        except Exception as e:
            print(f"  ⚠ Error for {phone}: {e}")


# ── Print terminal report ─────────────────────────────────────────

def print_report(results):
    print("\n" + "=" * 65)
    print(f"  {RACE_FLAG}  {RACE_NAME}  |  Race {MATCHDAY}")
    print(f"  League: {LEAGUE_NAME}")
    print(f"  Rankings recalculated from live data")
    print("=" * 65)

    for t in results:
        arrow = rank_change_arrow(t["rank_change"])
        print(f"\n{'─'*65}")
        print(f"  #{t['overall_rank']} (GD #{t['live_gd_rank']})  "
              f"{t['team_name']}  ({t['user_name']})  {arrow}")
        print(f"  Total pts : {t['total_points']}  |  This race: {t['gd_points']}")
        print(f"  Budget    : ${t['budget_remaining']:.1f}m  |  "
              f"Team value: ${t['team_value']:.1f}m")
        cards_str = ", ".join(t["cards_used"]) if t["cards_used"] else "None"
        print(f"  Cards used: {cards_str}")
        print(f"\n  {'Type':<13} {'Name':<28} {'GD Pts':>6}  Breakdown")
        print(f"  {'─'*13} {'─'*28} {'─'*6}  {'─'*28}")
        for p in t["picks"]:
            pts_str = str(p["points_gd"]) if p["points_gd"] is not None else "?"
            cap     = " ★CAP"  if p["is_captain"] else \
                      " ★MEGA" if p["is_megacap"]  else ""
            name    = p["name"] + cap
            bd      = " | ".join(p["breakdown"]) if p["breakdown"] else ""
            print(f"  {p['type']:<13} {name:<34} {pts_str:>4}  {bd}")


# ── Main ──────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  F1 FANTASY LEAGUE — FETCH STANDINGS")
    print("=" * 65)
    print(f"  Race     : {RACE_FLAG} {RACE_NAME} (Race {MATCHDAY})")
    print(f"  League   : {LEAGUE_NAME} ({LEAGUE_ID})")
    print(f"  Sheets   : {'✅ Enabled' if ENABLE_SHEETS else '⏭ Disabled'}")
    print(f"  Save to  : {SAVE_FOLDER}")

    if not COOKIE:
        print("\n⚠️  ERROR: Cookie not loaded! Check cookie.txt in the script folder.")
        return

    results = fetch_all()
    print_report(results)

    print("\n[5/5] Saving + syncing...")
    save_files(results)
    update_google_sheet(results)

    print("\n✅ All done!")


if __name__ == "__main__":
    main()


# Allow calling as module from run_all.py
run = main
