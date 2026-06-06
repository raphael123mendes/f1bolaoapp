"""
Fantasy F1 2026 – Excel Dataset Builder
========================================
Run after any file update in your Google Drive folder.

Usage:
  python build_fantasy_f1.py

Outputs (all written to G:\\My Drive\\FantasyF1-26):
  fantasy_f1_dataset.xlsx   – 5 tabs: Races, Prices, Results, Picks, Breakdowns
  races.json
  prices.json
  results.json
  picks.json
  breakdowns.json
"""

import json, re, sys, base64, os
import requests
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FOLDER       = Path(r"G:\My Drive\FantasyF1-26")
TOKEN_FILE   = FOLDER / "github_token.txt"

# ── Style helpers ──────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", start_color="1F3864")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ROW_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="EAF0FB")

def style_header(ws, col_widths):
    for col_idx, width in enumerate(col_widths, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

def style_rows(ws, cancelled_col=None):
    for row in ws.iter_rows(min_row=2):
        cancelled = cancelled_col and row[cancelled_col - 1].value in (True, "TRUE", "True", 1)
        for cell in row:
            cell.font = ROW_FONT
            if cancelled:
                cell.fill = PatternFill("solid", start_color="FFE0E0")
            elif cell.row % 2 == 0:
                cell.fill = ALT_FILL

def autofilter(ws):
    ws.auto_filter.ref = ws.dimensions

# ── Helpers ────────────────────────────────────────────────────────────────────
def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

BREAKDOWN_RE = re.compile(r'^(.+?):\s*([+-]?\d+)$')

PRIZES = {1: 19.25, 2: 11.55, 3: 7.70}

def calc_prizes(results_for_race):
    from collections import defaultdict
    by_rank = defaultdict(list)
    for r in results_for_race:
        by_rank[r["race_rank"]].append(r["user_name"])
    out = {}
    for rank in [1, 2, 3]:
        users = by_rank[rank]
        if not users:
            continue
        pool = sum(PRIZES[p] for p in range(rank, rank + len(users)) if p in PRIZES)
        share = round(pool / len(users), 4)
        for user in users:
            out[user] = share
    return out

# ── Data builder ───────────────────────────────────────────────────────────────
def build_data(schedule, prices_data, race_files):
    """Build all four datasets from source files. Returns dict of lists."""

    # ── Races: one row per round ───────────────────────────────────────────────
    races = []
    for r in schedule["rounds"]:
        race_session = next((s for s in r["sessions"] if s["session_type"] == "race"), None)
        races.append({
            "round":               r["round"],
            "race_name":           r["name"],
            "location":            r["location"],
            "venue":               r["venue"],
            "race_date":           race_session["utc_start"][:10] if race_session else "",
            "sprint_weekend":      r.get("sprint_weekend", False),
            "cancelled":           r.get("cancelled", False),
            "cancellation_reason": r.get("cancellation_reason", ""),
            "notes":               r.get("notes", ""),
        })

    # ── Prices: one row per player × race ─────────────────────────────────────
    prices = prices_data["prices"]

    # ── Price lookup for budget calc ──────────────────────────────────────────
    price_lookup = {
        (str(p["race_number"]), str(p["player_id"])): p
        for p in prices
    }

    # ── Results, Picks, Breakdowns from race files ─────────────────────────────
    results    = []
    picks      = []
    breakdowns = []

    # First pass: collect all race entries grouped by race_number
    # so we can compute rank_change and prizes per race
    all_races_entries = {}   # race_number -> list of raw entries
    for path in sorted(race_files):
        data = load_json(path)
        if data:
            all_races_entries[data[0]["race_number"]] = data

    for race_number in sorted(all_races_entries):
        race_data    = all_races_entries[race_number]
        prev_number  = race_number - 1
        prev_data    = all_races_entries.get(prev_number, [])
        prev_rank    = {e["user_name"]: e["rank"] for e in prev_data}

        # Build partial results list for prize calc (uses race_rank = live_gd_rank)
        partial = [{"user_name": e["user_name"], "race_rank": e["live_gd_rank"]} for e in race_data]
        prizes  = calc_prizes(partial)

        for e in race_data:
            budget_remaining = e["budget_remaining"]
            team_value       = e["team_value"]
            current_budget   = round(team_value + budget_remaining, 2)

            # Compute total new budget across all picks
            new_budget = current_budget
            for pick in e.get("picks", []):
                key = (str(e["race_number"]), str(pick["id"]))
                pdata = price_lookup.get(key, {})
                pt = pdata.get("price_this_race")
                pn = pdata.get("price_next_race")
                if pt is not None and pn is not None:
                    new_budget += (pn - pt)
            new_budget = round(new_budget, 2)

            # rank_change: positive = moved up, negative = dropped
            # Race 1 has no previous so 0
            if e["user_name"] in prev_rank:
                rank_change = prev_rank[e["user_name"]] - e["rank"]
            else:
                rank_change = 0

            # ── Results row (one per user × race) ─────────────────────────────
            results.append({
                "race_number":      e["race_number"],
                "race_name":        e["race_name"],
                "race_flag":        e.get("race_flag", ""),
                "race_date":        e["race_date"],
                "season_rank":      e["rank"],
                "race_rank":        e["live_gd_rank"],
                "rank_change":      rank_change,
                "team_name":        e["team_name"],
                "user_name":        e["user_name"],
                "season_points":    e["total_points"],
                "race_points":      e["gd_points"],
                "budget_remaining": budget_remaining,
                "team_value":       team_value,
                "cards_used":       "; ".join(e.get("cards_used", [])),
                "current_budget":   current_budget,
                "new_budget":       new_budget,
                "race_prize":       prizes.get(e["user_name"], 0),
            })

            # ── Picks rows (one per pick) ──────────────────────────────────────
            for pick in e.get("picks", []):
                key   = (str(e["race_number"]), str(pick["id"]))
                pdata = price_lookup.get(key, {})
                picks.append({
                    "race_number":    e["race_number"],
                    "race_name":      e["race_name"],
                    "race_flag":      e.get("race_flag", ""),
                    "race_date":      e["race_date"],
                    "season_rank":    e["rank"],
                    "race_rank":      e["live_gd_rank"],
                    "team_name":      e["team_name"],
                    "user_name":      e["user_name"],
                    "cards_used":     "; ".join(e.get("cards_used", [])),
                    "pick_id":        pick["id"],
                    "pick_name":      pick["name"],
                    "pick_type":      pick["type"],
                    "is_captain":     pick["is_captain"],
                    "is_megacap":     pick.get("is_megacap", False),
                    "pick_points_gd": pick["points_gd"],
                    "price_this_race": pdata.get("price_this_race"),
                    "price_next_race": pdata.get("price_next_race"),
                })

                # ── Breakdowns rows (one per breakdown line) ───────────────────
                for bd in pick.get("breakdown", []):
                    m = BREAKDOWN_RE.match(bd.strip())
                    breakdowns.append({
                        "race_number":      e["race_number"],
                        "race_name":        e["race_name"],
                        "race_flag":        e.get("race_flag", ""),
                        "race_date":        e["race_date"],
                        "season_rank":      e["rank"],
                        "race_rank":        e["live_gd_rank"],
                        "team_name":        e["team_name"],
                        "user_name":        e["user_name"],
                        "pick_id":          pick["id"],
                        "pick_name":        pick["name"],
                        "pick_type":        pick["type"],
                        "is_captain":       pick["is_captain"],
                        "pick_points_gd":   pick["points_gd"],
                        "breakdown_category": m.group(1).strip() if m else bd.strip(),
                        "breakdown_points":   int(m.group(2)) if m else None,
                    })

    return {"races": races, "prices": prices, "results": results,
            "picks": picks, "breakdowns": breakdowns}


# ── Excel writers ──────────────────────────────────────────────────────────────
def write_races(wb, data):
    ws = wb.create_sheet("Races")
    ws.append(["Round", "Race Name", "Location", "Venue", "Race Date",
                "Sprint Weekend", "Cancelled", "Cancellation Reason", "Notes"])
    for r in data:
        ws.append([r["round"], r["race_name"], r["location"], r["venue"],
                   r["race_date"], r["sprint_weekend"], r["cancelled"],
                   r["cancellation_reason"], r["notes"]])
    style_header(ws, [7, 30, 25, 32, 12, 14, 10, 22, 35])
    style_rows(ws, cancelled_col=7)
    autofilter(ws)

def write_prices(wb, data):
    ws = wb.create_sheet("Prices")
    ws.append(["Race Number", "Race Name", "Race Flag", "Race Date",
               "Next Race Number", "Next Race Name",
               "Player ID", "Player Name", "Type",
               "Price This Race", "Price Prev Race", "Price Next Race",
               "Price Change", "Price Change %",
               "Gameday Points", "Total Points",
               "Season High", "Season Low",
               "Pct vs High", "Pct vs Low",
               "Fetched At"])
    for p in data:
        ws.append([p["race_number"], p["race_name"], p.get("race_flag",""),
                   p["race_date"], p.get("next_race_number"), p.get("next_race_name"),
                   p["player_id"], p["player_name"], p["type"],
                   p["price_this_race"], p["price_prev_race"], p.get("price_next_race"),
                   p["price_change"], p["price_change_pct"],
                   p.get("gameday_points"), p.get("total_points"),
                   p.get("season_high"), p.get("season_low"),
                   p.get("pct_vs_high"), p.get("pct_vs_low"),
                   p.get("fetched_at","")])
    style_header(ws, [13, 28, 10, 12, 15, 24, 10, 22, 12, 16, 16, 16, 14, 14, 15, 13, 13, 12, 12, 12, 22])
    style_rows(ws)
    autofilter(ws)

def write_results(wb, data):
    ws = wb.create_sheet("Results")
    ws.append(["Race Number", "Race Name", "Race Flag", "Race Date",
               "Season Rank", "Race Rank", "Rank Change",
               "Team Name", "User Name",
               "Season Points", "Race Points",
               "Budget Remaining", "Team Value", "Cards Used",
               "Current Budget", "New Budget", "Race Prize"])
    for r in data:
        ws.append([r["race_number"], r["race_name"], r["race_flag"], r["race_date"],
                   r["season_rank"], r["race_rank"], r["rank_change"],
                   r["team_name"], r["user_name"],
                   r["season_points"], r["race_points"],
                   r["budget_remaining"], r["team_value"], r["cards_used"],
                   r["current_budget"], r["new_budget"], r["race_prize"]])
    style_header(ws, [12, 28, 10, 12, 11, 10, 12, 22, 22, 13, 12, 16, 11, 28, 15, 12, 11])
    style_rows(ws)
    autofilter(ws)

def write_picks(wb, data):
    ws = wb.create_sheet("Picks")
    ws.append(["Race Number", "Race Name", "Race Flag", "Race Date",
               "Season Rank", "Race Rank", "Team Name", "User Name", "Cards Used",
               "Pick ID", "Pick Name", "Pick Type", "Is Captain", "Is Megacap",
               "Pick Points GD", "Price This Race", "Price Next Race"])
    for p in data:
        ws.append([p["race_number"], p["race_name"], p["race_flag"], p["race_date"],
                   p["season_rank"], p["race_rank"], p["team_name"], p["user_name"],
                   p["cards_used"],
                   p["pick_id"], p["pick_name"], p["pick_type"],
                   p["is_captain"], p["is_megacap"], p["pick_points_gd"],
                   p["price_this_race"], p["price_next_race"]])
    style_header(ws, [12, 28, 10, 12, 11, 10, 22, 22, 28,
                      10, 22, 14, 10, 10, 14, 15, 15])
    style_rows(ws)
    autofilter(ws)

def write_breakdowns(wb, data):
    ws = wb.create_sheet("Breakdowns")
    ws.append(["Race Number", "Race Name", "Race Flag", "Race Date",
               "Season Rank", "Race Rank", "Team Name", "User Name",
               "Pick ID", "Pick Name", "Pick Type", "Is Captain",
               "Pick Points GD", "Breakdown Category", "Breakdown Points"])
    for b in data:
        ws.append([b["race_number"], b["race_name"], b["race_flag"], b["race_date"],
                   b["season_rank"], b["race_rank"], b["team_name"], b["user_name"],
                   b["pick_id"], b["pick_name"], b["pick_type"], b["is_captain"],
                   b["pick_points_gd"], b["breakdown_category"], b["breakdown_points"]])
    style_header(ws, [12, 28, 10, 12, 11, 10, 22, 22, 10, 22, 14, 10, 14, 26, 16])
    style_rows(ws)
    autofilter(ws)


# ── GitHub push ───────────────────────────────────────────────────────────────
GITHUB_REPO  = "raphael123mendes/f1bolaoapp"
GITHUB_FILES = ["races.json", "prices.json", "results.json", "picks.json", "breakdowns.json"]

def push_to_github(folder: Path, token: str):
    """Push JSON output files to GitHub repo root via the Contents API."""
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github+json",
    }
    base_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents"
    success, failed = [], []

    for filename in GITHUB_FILES:
        filepath = folder / filename
        if not filepath.exists():
            print(f"  [SKIP] {filename} not found")
            continue

        with open(filepath, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()

        # Check if file already exists (need its SHA to update)
        url = f"{base_url}/{filename}"
        r = requests.get(url, headers=headers)
        sha = r.json().get("sha") if r.status_code == 200 else None

        payload = {
            "message": f"Update {filename}",
            "content": encoded,
        }
        if sha:
            payload["sha"] = sha

        r = requests.put(url, headers=headers, json=payload)
        if r.status_code in (200, 201):
            action = "updated" if sha else "created"
            print(f"  ✓  {filename} {action}")
            success.append(filename)
        else:
            print(f"  ✗  {filename} failed: {r.status_code} {r.json().get('message','')}")
            failed.append(filename)

    print(f"\nGitHub: {len(success)} pushed, {len(failed)} failed → {GITHUB_REPO}")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else FOLDER
    if not folder.exists():
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)

    schedule_path = folder / "f1_2026_schedule.json"
    prices_path   = folder / "prices_master.json"
    race_files    = sorted(folder.glob("race_*.json"))

    for p in [schedule_path, prices_path]:
        if not p.exists():
            print(f"ERROR: Missing file: {p.name}")
            sys.exit(1)

    print(f"Fantasy F1 Dataset Builder  |  folder: {folder}")
    print(f"Race files: {len(race_files)}")

    data = build_data(load_json(schedule_path), load_json(prices_path), race_files)

    # ── Excel ──────────────────────────────────────────────────────────────────
    print("\nBuilding Excel...")
    wb = Workbook()
    wb.remove(wb.active)
    write_races(wb, data["races"])
    write_prices(wb, data["prices"])
    write_results(wb, data["results"])
    write_picks(wb, data["picks"])
    write_breakdowns(wb, data["breakdowns"])
    wb.save(str(folder / "fantasy_f1_dataset.xlsx"))

    # ── JSON exports ───────────────────────────────────────────────────────────
    print("Exporting JSON files...")
    for name, rows in data.items():
        save_json(folder / f"{name}.json", rows)

    print("\nDone! Outputs:")
    for name, rows in data.items():
        print(f"  {name+'.json':<20} {len(rows):>4} rows")
    print(f"  fantasy_f1_dataset.xlsx  (5 tabs)")

    # ── GitHub push ────────────────────────────────────────────────────────────
    token = ""
    token_file = folder / "github_token.txt"
    if token_file.exists():
        token = token_file.read_text(encoding="utf-8").strip()
    if token:
        print("\nPushing JSON files to GitHub...")
        push_to_github(folder, token)
    else:
        print(f"\nSkipping GitHub push — no token found.")
        print(f"  Create {folder / 'github_token.txt'} with your GitHub PAT to enable auto-push.")

if __name__ == "__main__":
    main()
